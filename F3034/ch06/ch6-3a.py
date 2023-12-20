from openpyxl import load_workbook

wb = load_workbook("水果2.xlsx")
ws = wb.create_sheet(title="中區", index=1)

ws["A1"] = "香蕉"
ws["A2"] = "1"
ws["A3"] = "4"
wb.save("水果3.xlsx ")
wb.close()

