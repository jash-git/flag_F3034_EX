from openpyxl import load_workbook

wb = load_workbook("營業額.xlsx")
ws = wb['工作表1']
for col in ws.iter_cols():
    for cell in col:
        print(cell.value, end=" ")
    print()  
print("--------------------")
for col in ws.iter_cols(values_only=True):
    print(col)  
print("--------------------")
for col in ws.iter_cols(min_row=1,
                        max_row=2,
                        min_col=1,
                        max_col=3):
    for cell in col:
        print(cell.value, end=" ")
    print()     
wb.close()
