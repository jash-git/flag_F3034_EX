from openpyxl import load_workbook

wb = load_workbook("ch5-2.xlsx")
ws = wb.active
ws.title = "姓名"
ws["C1"] = "陳會安"
wb.save("ch5-2a.xlsx")

