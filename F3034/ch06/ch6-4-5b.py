from openpyxl import load_workbook

wb = load_workbook("Excel水果資料2.xlsx")
ws = wb.active
ws.unmerge_cells("A2:C2")
ws.unmerge_cells("A3:C3")
wb.save("Excel水果資料3.xlsx")   
wb.close()


