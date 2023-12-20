from openpyxl import load_workbook

wb = load_workbook("水果3.xlsx")
print(wb.sheetnames)
print(len(wb.sheetnames))
print(wb.sheetnames[0])
print(wb.sheetnames[1])
for ws_name in wb.sheetnames:
    print(ws_name)
wb.close()
