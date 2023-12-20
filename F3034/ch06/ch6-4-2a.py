from openpyxl import load_workbook

wb = load_workbook("營業額.xlsx")
ws = wb['工作表1']
cells = ws["A1:D3"]
for i1, i2, i3, i4 in cells:
    print(i1.value, i2.value,
          i3.value, i4.value)
wb.close()
