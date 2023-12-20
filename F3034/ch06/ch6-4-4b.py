from openpyxl import load_workbook

wb = load_workbook("Excel水果資料.xlsx")
ws = wb.active

def print_rows():
    for row in ws.iter_rows(values_only=True):
        print(row)
        
print_rows()
print("--------------------")
ws.insert_rows(idx=1)
print_rows()
print("--------------------")
ws.insert_rows(idx=3, amount=2)
print_rows()
print("--------------------")
wb.close()
