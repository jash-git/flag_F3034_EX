from openpyxl import load_workbook

wb = load_workbook("Excel水果資料.xlsx")
ws = wb.active

def print_rows():
    for row in ws.iter_rows(values_only=True):
        print(row)
        
print_rows()
print("--------------------")
ws.delete_rows(idx=1, amount=1)
print_rows()
print("--------------------")
ws.delete_cols(idx=3, amount=1)
print_rows()
print("--------------------")
wb.close()
