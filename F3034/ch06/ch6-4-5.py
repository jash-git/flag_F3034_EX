from openpyxl import load_workbook

wb = load_workbook("Excel水果資料.xlsx")
ws = wb.active

def print_rows():
    for row in ws.iter_rows(values_only=True):
        print(row)
        
print_rows()
print("--------------------")
ws.move_range("A1:C4", rows=1, cols=1)
print_rows()
print("--------------------")
ws.move_range("B2:D5", rows=-1, cols=-1)
print_rows()
wb.close()
