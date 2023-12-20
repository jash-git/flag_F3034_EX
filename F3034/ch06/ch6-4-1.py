from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "東區"
ws["A1"] = "香蕉"
ws["B1"] = "蘋果"
ws["C1"] = "西瓜"
ws.cell(row=2, column=1).value = 1
ws.cell(row=2, column=2).value = 3
ws.cell(row=2, column=3).value = 2
wb.save("水果資料.xlsx")   
wb.close()
