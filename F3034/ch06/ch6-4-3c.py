from openpyxl import load_workbook

wb = load_workbook("營業額.xlsx")
ws = wb['工作表1']
for row in ws.iter_rows():
    for cell in row:
        print(cell.value, end=" ")
    print()  
print("--------------------")
for row in ws.iter_rows(values_only=True):
    print(row)  
print("--------------------")
for row in ws.iter_rows(min_row=1,
                        max_row=2,
                        min_col=1,
                        max_col=3):
    for cell in row:
        print(cell.value, end=" ")
    print()     
wb.close()
