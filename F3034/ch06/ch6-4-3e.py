from openpyxl import load_workbook

def get_rows(ws, min_row, max_row ,min_col , max_col):
    rows_range = ws.iter_rows(max_row=max_row,min_row=min_row,
                              max_col=max_col,min_col=min_col)
    rows = []
    for row in rows_range:
        columns = []
        for column in row:
            columns.append(column.value)
        rows.append(columns)
    return rows

wb = load_workbook("營業額.xlsx")
ws = wb['工作表1']
rows = get_rows(ws, 2, 4, 1, 3)
for row in rows:
    print(row)
wb.close()
