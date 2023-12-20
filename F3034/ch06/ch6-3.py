from openpyxl import load_workbook

wb = load_workbook("水果.xlsx")
ws = wb.create_sheet(title="南區")

ws["A1"] = "蘋果"
ws["A2"] = "2"
ws["A3"] = "3"
wb.save("水果2.xlsx ")
wb.close()
