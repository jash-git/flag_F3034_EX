from openpyxl import load_workbook

wb = load_workbook("營業額.xlsx", read_only=True)
print(wb.sheetnames)
ws = wb.active
print("工作表名稱: ", ws.title)
ws["D2"] = 3500
wb.close()


