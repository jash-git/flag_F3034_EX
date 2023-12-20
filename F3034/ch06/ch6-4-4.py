from openpyxl import load_workbook

wb = load_workbook("營業額.xlsx")
ws = wb['工作表1']
col = ws["C"]
for cell in col:
    print(cell.value, end=" ")
print()
row = ws[3]
for cell in row:
    print(cell.value, end=" ")
print()
print("--------------------")
cols = ws["C:D"]
for col in cols:
    for cell in col:
        print(cell.value, end=" ")
    print()  
print("--------------------")
rows = ws[3:4]
for row in rows:
    for cell in row:
        print(cell.value, end=" ")
    print() 
wb.close()
