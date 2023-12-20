from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "測試值"
ws["A1"] = "Hello"
ws["B1"] = "World"
wb.save("ch5-2.xlsx")


