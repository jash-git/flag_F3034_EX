from openpyxl import load_workbook

wb = load_workbook("Excel水果資料.xlsx")
ws = wb.active
ws.merge_cells("A2:C2")
ws.merge_cells("A3:C3")
wb.save("Excel水果資料2.xlsx")   
wb.close()

