from openpyxl import load_workbook

wb = load_workbook("水果.xlsx")
ws = wb.active

print("工作表名稱:", ws.title)
print("最小欄索引:", ws.min_column)
print("最大欄索引:", ws.max_column)
print("最小列索引:", ws.min_row)
print("最大列索引:", ws.max_row)
print("工作表尺寸:", ws.dimensions)

from openpyxl.utils import get_column_letter, column_index_from_string

print("索引1的字母:", get_column_letter(1))
print("字母C的索引:", column_index_from_string("C"))
print("最小欄字母:", get_column_letter(ws.min_column))
print("最大欄字母:", get_column_letter(ws.max_column))
wb.close()
