from openpyxl import load_workbook

wb = load_workbook("水果3.xlsx")
ws = wb.active
print("目前工作表名稱:", ws.title)
wb.active = 2
ws = wb.active
print("目前工作表名稱:", ws.title)
print(wb.sheetnames)
ws2 = wb["工作表1"]
print("切換至工作表1:", ws2.title)
ws3 = wb["中區"]
print("切換至中區:", ws3.title)
wb.close()
