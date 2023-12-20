from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

wb = load_workbook("2330TW.xlsx")
ws = wb.active

pattern_fill = PatternFill(start_color="00EE1111",
                           end_color="00EE1111")
pattern_fill2 = PatternFill(start_color="0066FF66",
                            end_color="0066FF66")
diff_style = DifferentialStyle(fill=pattern_fill)
diff_style2 = DifferentialStyle(fill=pattern_fill2)
rule = Rule(type="cellIs", 
            operator="lessThan", 
            formula=[264],
            dxf=diff_style)
rule2 = Rule(type="cellIs")
rule2.operator="greaterThan"
rule2.formula=[266]
rule2.dxf=diff_style2
ws.conditional_formatting.add("B2:F19", rule)
ws.conditional_formatting.add("B2:F19", rule2)
wb.save("ConditionalFormatting2.xlsx")
wb.close()