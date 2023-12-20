from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

wb = load_workbook("2330TW.xlsx")
ws = wb.active

pattern_fill = PatternFill(start_color="00EE1111",
                           end_color="00EE1111")
diff_style = DifferentialStyle(fill=pattern_fill)
rule = Rule(type="cellIs", 
            operator="lessThan", 
            formula=[264],
            dxf=diff_style)

ws.conditional_formatting.add("B2:F19", rule)
wb.save("ConditionalFormatting.xlsx")
wb.close()
