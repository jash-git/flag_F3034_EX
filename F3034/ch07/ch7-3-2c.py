from openpyxl import Workbook
 
wb = Workbook()
ws = wb.active
ws.title = "QUOTIENT"
 
data = [["數字"], [100], [200], [300], [400]]
for i in data:
    ws.append(i)
 
ws["C1"] = "整除8="
for i in range(2, 6):
    ws["C"+str(i)] = "= QUOTIENT(A"+ str(i) +", 8)"
 
wb.save("ExcelQuotient.xlsx")
wb.close()