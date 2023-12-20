from openpyxl import load_workbook

wb = load_workbook("Excel資料.xlsx")
ws = wb.active
ws["D1"] = "小計"
ws["D2"] = "=SUM(A2:C2)"
ws["D3"] = "=SUM(A3:C3)"
wb.save("Excel資料2.xlsx")   
wb.close()


