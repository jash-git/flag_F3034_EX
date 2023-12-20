from openpyxl import Workbook
 
wb = Workbook()
ws = wb.active
ws.title = "MOD"
 
data = [["數字"], [100], [200], [300], [400]]
for i in data:
    ws.append(i)
 
ws["C1"] = "除8餘數="
for i in range(2, 6):
    ws["C"+str(i)] = "= MOD(A"+ str(i) +", 8)"
 
wb.save("ExcelModulus.xlsx")
wb.close()