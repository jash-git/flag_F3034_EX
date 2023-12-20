from openpyxl import Workbook
from openpyxl.styles import Alignment
  
wb = Workbook()
ws = wb.active
ws.title = "對齊樣式"
# 建立對齊物件
left_aligned = Alignment(horizontal="left")
right_aligned = Alignment(horizontal="right")
center_aligned = Alignment(horizontal="center")
# 靠左
ws.column_dimensions["A"].width = 20
ws["A1"].alignment = left_aligned
ws.cell(row=1, column=1).value = "Python"
# 置中
ws.column_dimensions["B"].width = 20
ws["B2"].alignment = center_aligned
ws.cell(row=2, column=2).value = "Excel"
# 靠右
ws.column_dimensions["C"].width = 20
ws["C3"].alignment = right_aligned
ws.cell(row=3, column=3).value = "Python"
wb.save("Excel儲存格對齊樣式.xlsx")
wb.close()