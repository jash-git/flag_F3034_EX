from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.styles import colors, Font, Alignment, Side, Border

wb = load_workbook("營業額.xlsx")
ws = wb.active
# 建立命名樣式company
company_column = NamedStyle(name="company")
company_column.font = Font(bold=True, size=20, 
                           color=colors.COLOR_INDEX[4])
company_column.alignment = Alignment(horizontal="right")
company_column.border = Border(right=Side(border_style="dotted", 
                               color=colors.COLOR_INDEX[2])) 
# 建立命名樣式header
header_row = NamedStyle(name="header")
header_row.font = Font(italic=True, bold=True)
header_row.alignment = Alignment(horizontal="center", 
                                 vertical="center")
header_row.border = Border(bottom=Side(border_style="thick")) 
# 替整欄儲存格套用company樣式
for cell in ws["A"]:
    cell.style = company_column
ws.column_dimensions["A"].width = 30
# 替整列儲存格套用header樣式
for cell in ws[1]:
    cell.style = header_row 
wb.save("營業額_Styles.xlsx") 
wb.close()
