from openpyxl import Workbook
from openpyxl.styles import colors, Font, Alignment, Side, Border
  
wb = Workbook()
ws = wb.active
ws.title = "框線樣式"
# 建立Side和Border物件
double_side = Side(color=colors.BLUE, 
                   border_style="double")
dashed_side = Side(style="dashed")
cell_border = Border(top=double_side,
                     bottom=double_side,
                     left=dashed_side,
                     right=dashed_side)
center_aligned = Alignment(horizontal="center",
                           vertical="center")
red_bold_font = Font(size=24, bold=True,
                     color=colors.COLOR_INDEX[2])
ws["B2"] = "Hello"
ws.column_dimensions["B"].width = 20
ws.row_dimensions[2].height = 70
ws["B2"].font = red_bold_font
ws["B2"].alignment = center_aligned
ws["B2"].border = cell_border
wb.save("Excel儲存格框線樣式.xlsx")
wb.close()