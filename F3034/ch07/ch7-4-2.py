from openpyxl import Workbook, load_workbook 

wb_output = Workbook()
ws_output = wb_output.active
ws_output["A1"].value = "成員"
ws_output["B1"].value = "數量"
ws_output["C1"].value = "金額"
ws_output["D1"].value = "月份"
ws_output["E1"].value = "數量"
ws_output["F1"].value = "金額"

wb = load_workbook("收支流水帳清單.xlsx")
ws = wb.active
data = {}
for row in range(2, ws.max_row + 1):
    member = ws["A" + str(row)].value
    month  = ws["B" + str(row)].value
    amount = ws["D" + str(row)].value
    
    if member not in data:
        data[member] = {"name": member,
                        "quantity": 0,
                        "amount": 0 }
    if month not in data[member]:
        data[member][month] = {"name": month ,
                               "quantity": 0,
                               "amount": 0 }
    data[member][month]["quantity"] += 1
    data[member][month]["amount"] += int(amount)
    data[member]["quantity"] += 1
    data[member]["amount"] += int(amount)

curr_row = 2
for member_data in data.values():
    ws_output["A"+str(curr_row)].value = member_data["name"]
    ws_output["B"+str(curr_row)].value = member_data["quantity"]
    ws_output["C"+str(curr_row)].value = member_data["amount"]
    for month_data in member_data.values():
        if isinstance(month_data, dict):
            ws_output["D"+str(curr_row)].value = month_data["name"]
            ws_output["E"+str(curr_row)].value = month_data["quantity"]
            ws_output["F"+str(curr_row)].value = month_data["amount"]
            curr_row = curr_row + 1 

curr_row = curr_row + 1 
ws_output["D"+str(curr_row)].value = "合計"
ws_output["E"+str(curr_row)].value = "=SUM(E2:E"+str(curr_row-1)+")"
ws_output["F"+str(curr_row)].value = "=SUM(F2:F"+str(curr_row-1)+")"

wb_output.save("收支流水帳統計_月份.xlsx")
wb.close()