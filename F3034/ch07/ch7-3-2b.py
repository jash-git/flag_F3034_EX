from openpyxl import Workbook
 
wb = Workbook()
ws = wb.active
ws.title = "AVERAGE"
 
data = [[100], [200], [300], [400]]
for i in data:
    ws.append(i)
 
ws["A6"] = "平均="
ws["B6"] = "= AVERAGE(A1:A4)"
wb.save("ExcelAverage.xlsx")
wb.close()