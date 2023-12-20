from openpyxl import Workbook
from openpyxl.styles import Font
  
wb = Workbook()
ws = wb.active
ws.title = "字型樣式"
# 指定字型尺寸
ws.cell(row=1, column=1).font = Font(size=24)
ws.cell(row=1, column=1).value = "Python X Excel"
# 指定斜體字
ws.cell(row=2, column=2).font = Font(size=24, italic=True)
ws.cell(row=2, column=2).value = "Python X Excel"
# 指定粗體字
ws.cell(row=3, column=3).font = Font(size=24, bold=True)
ws.cell(row=3, column=3).value = "Python X Excel"
# 指定底線字
ws.cell(row=4, column=4).font = Font(size=24, underline="double")
ws.cell(row=4, column=4).value = "Python X Excel"
# 指定哪一種字型
ws.cell(row=5, column=5).font = Font(size=24, name='Times New Roman')
ws.cell(row=5, column=5).value = "Python X Excel"  
wb.save("Excel儲存格字型樣式.xlsx")
wb.close()