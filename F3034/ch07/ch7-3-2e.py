from openpyxl import Workbook
 
wb = Workbook()
ws = wb.active
ws.title = "COUNT"
 
data = [[100], [200], [300], [400]]
for i in data:
    ws.append(i)
 
ws["A6"] = "計數="
ws["B6"] = "= COUNT(A1:A4)" 
wb.save("ExcelCount.xlsx")
wb.close()