from openpyxl import Workbook
from openpyxl.styles import PatternFill, GradientFill
  
wb = Workbook()
ws = wb.active
ws.title = "填充樣式"
# 建立填充物件
pattern_fill = PatternFill(fill_type="solid",
                           fgColor="0066FF66")
gradient_fill = GradientFill(stop=("00FFFFFF",
                                   "0099CCFF",
                                   "00000000"))
# 型樣填充
ws.column_dimensions["A"].width = 20
ws["A1"].fill = pattern_fill
ws.cell(row=1, column=1).value = "Python"
# 漸層填充
ws.column_dimensions["B"].width = 20
ws["B2"].fill = gradient_fill
ws.cell(row=2, column=2).value = "Excel"
wb.save("Excel儲存格填充樣式.xlsx")
wb.close()