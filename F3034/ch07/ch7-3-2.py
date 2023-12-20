from openpyxl import Workbook
 
wb = Workbook()
ws = wb.active
ws.title = "SUM"
 
data = [[100], [200], [300], [400]]
for i in data:
    ws.append(i)
 
ws["A6"] = "總和="
ws["B6"] = "= SUM(A1:A4)"
wb.save("ExcelSum.xlsx")
wb.close()