from openpyxl import Workbook, load_workbook 

wb_output = Workbook()
ws_output = wb_output.active
ws_output["A1"].value = "成員"
ws_output["B1"].value = "數量"
ws_output["C1"].value = "金額"
ws_output["D1"].value = "項目"
ws_output["E1"].value = "數量"
ws_output["F1"].value = "金額"

wb = load_workbook("收支流水帳清單.xlsx")
ws = wb.active
data = {}
for row in range(2, ws.max_row + 1):
    member = ws["A" + str(row)].value
    item   = ws["C" + str(row)].value
    amount = ws["D" + str(row)].value
    
    data.setdefault(member,
                    {"name": member,
                     "quantity": 0,
                     "amount": 0
                    })
    data[member].setdefault(item, {"name": item ,
                                    "quantity": 0,
                                    "amount": 0
                                   })    
    data[member][item]["quantity"] += 1
    data[member][item]["amount"] += int(amount)
    data[member]["quantity"] += 1
    data[member]["amount"] += int(amount)

curr_row = 2
for member_data in data.values():
    ws_output["A"+str(curr_row)].value = member_data["name"]
    ws_output["B"+str(curr_row)].value = member_data["quantity"]
    ws_output["C"+str(curr_row)].value = member_data["amount"]
    for item_data in member_data.values():
        if isinstance(item_data, dict):
            ws_output["D"+str(curr_row)].value = item_data["name"]
            ws_output["E"+str(curr_row)].value = item_data["quantity"]
            ws_output["F"+str(curr_row)].value = item_data["amount"]
            curr_row = curr_row + 1 

curr_row = curr_row + 1 
ws_output["D"+str(curr_row)].value = "合計"
ws_output["E"+str(curr_row)].value = "=SUM(E2:E"+str(curr_row-1)+")"
ws_output["F"+str(curr_row)].value = "=SUM(F2:F"+str(curr_row-1)+")"

wb_output.save("收支流水帳統計_項目.xlsx")
wb.close()