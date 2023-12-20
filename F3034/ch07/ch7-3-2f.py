from openpyxl import Workbook
 
wb = Workbook()
ws = wb.active
ws.title = "COUNTIF"
 
data = [[100], [200], [300], [400]]
for i in data:
    ws.append(i)
 
ws["A6"] = "條件計數="
ws["B6"] = "= COUNTIF(A1:A4, \">250\")"
wb.save("ExcelCountIF.xlsx")
wb.close()