from openpyxl import load_workbook

wb = load_workbook("Excel水果資料.xlsx")
ws = wb["東區"]
ws["D1"] = "小計"
for i in range(2, 5):
    pos = "D"+str(i)
    ws[pos] = "=SUM(A"+str(i)+":C"+str(i)+")"

wb.save("Excel水果資料_Sum.xlsx")   
wb.close()
