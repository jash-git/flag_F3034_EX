from pathlib import Path
from openpyxl import Workbook, load_workbook

wb_output = Workbook() 
ws_output = wb_output.active 
curr_row = 1
# 新增標題列
ws_output.cell(curr_row,1).value = "成員"
ws_output.cell(curr_row,2).value = "月份"
ws_output.cell(curr_row,3).value = "項目"
ws_output.cell(curr_row,4).value = "金額"
curr_row = curr_row + 1
# 彙整家庭收支流水帳
path = Path("家庭收支流水帳/")
for item in path.iterdir():
    if item.match("*.xlsx"):
        wb = load_workbook(item)
        ws = wb.active
        name = ws["B2"].value
        for row in range(4, ws.max_row + 1):
            ws_output.cell(curr_row,1).value = name   # 姓名
            ws_output.cell(curr_row,2).value = ws["A"+str(row)].value # 月份                    
            ws_output.cell(curr_row,3).value = ws["B"+str(row)].value # 項目 
            ws_output.cell(curr_row,4).value = ws["C"+str(row)].value # 金額
            curr_row = curr_row + 1
        wb.close()    

wb_output.save("收支流水帳清單.xlsx")
wb_output.close()
