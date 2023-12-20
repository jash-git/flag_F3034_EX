from openpyxl import Workbook
  
wb = Workbook()
ws = wb.active
ws.title = "儲存格尺寸"

ws.cell(row=1, column=1).value = "Hello"
ws.cell(row=2, column=2).value = "陳會安"
ws.cell(row=3, column=3).value = "Python" 
ws.row_dimensions[1].height = 70
ws.column_dimensions["B"].width = 20
ws.row_dimensions[3].height = 50
ws.column_dimensions["C"].width = 7
wb.save("Excel儲存格尺寸.xlsx")
wb.close()