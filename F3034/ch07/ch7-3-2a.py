from openpyxl import Workbook
 
wb = Workbook()
ws = wb.active
ws.title = "PRODUCT"
 
data = [[10], [20], [30], [40]]
for i in data:
    ws.append(i)
 
ws["A6"] = "乘積="
ws["B6"] = "= PRODUCT(A1:A4)"
wb.save("ExcelProduct.xlsx")
wb.close()