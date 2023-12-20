from openpyxl import load_workbook

wb = load_workbook("Excel水果資料.xlsx")
ws = wb["東區"]
cols =["A", "B", "C"]
for col in cols:
    pos = col+"6"
    ws[pos] = "=SUM("+col+"2:"+col+"4)"

wb.save("Excel水果資料_Sum2.xlsx")   
wb.close()
