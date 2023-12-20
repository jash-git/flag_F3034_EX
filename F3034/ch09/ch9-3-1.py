from openpyxl import load_workbook, Workbook
                    
excel_files = ["Ubike01.xlsx", "Ubike02.xlsx", "Ubike03.xlsx"]
out_sheet = []
for excel in excel_files:
    wb = load_workbook(excel)
    ws = wb.active    
    for row in ws.rows:
       value = []
       for cell in row:
           value.append(str(cell.value))
       out_sheet.append(value)
    wb.close()   
                  
out_wb = Workbook()
out_ws = out_wb.active
for y, row in enumerate(out_sheet):
    for x, cell in enumerate(row):
        out_ws.cell(row=y+1,
                    column=x+1,
                    value=out_sheet[y][x])
                    
out_wb.save("Joined_Ubike.xlsx")
out_wb.close()