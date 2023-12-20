import json
from openpyxl import load_workbook

wb = load_workbook("營業額.xlsx")
ws = wb.active

keys = list(item.value for item in ws[1])
print(keys)
sales_list = []
for i in range(1, ws.max_row):
    row = ws[i+1]
    item = {}
    for index, key in enumerate(keys):
        item[key] = row[index].value
    sales_list.append(item)
    
print(sales_list)
jsonfile = "營業額.json"
with open(jsonfile, 'w', encoding="utf8") as fp:
    json.dump(sales_list, fp, ensure_ascii=False)
wb.close()