import json
from openpyxl import Workbook

keys = []
wb = Workbook()
ws = wb.active

with open("Books.json", encoding="utf8") as f:
    json_data = json.load(f)
	
for i in range(len(json_data)):
    json_obj = json_data[i]
    if i == 0:
        keys = list(json_obj.keys())
        for k in range(len(keys)):
            ws.cell(row=1, column=(k+1)).value = keys[k]
    for j in range(len(keys)):
        ws.cell(row=(i+2), column=(j+1)).value = json_obj[keys[j]]

wb.save("Books.xlsx")
wb.close()
