from win32com.client import Dispatch
import os
from PIL import ImageGrab
from openpyxl import load_workbook
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Cm

template_file = "Word報告模版.docx"
output_file   = "各通路業積報告.docx"
output_img    = "各通路業積圖表.png"
excel_file    = "Excel統計圖表.xlsx"

app = Dispatch("Excel.Application")
app.Visible = 1
app.DisplayAlerts = 0
xlsx = app.Workbooks.Open(os.getcwd()+"/"+excel_file)
sheet = xlsx.Worksheets(1)
for x, chart in enumerate(sheet.Shapes):
    chart.Copy()
    image = ImageGrab.grabclipboard()
    image.save(output_img, "png")
    
xlsx.Close(False)
app.Quit()

wb = load_workbook(excel_file)
ws = wb.active
sales = []
for row in ws.iter_rows(min_row=2, values_only=True):
    data = { "internet": row[1],
             "store": row[2],
             "direct": row[3] }
    sales.append(data)
wb.close()

doc = DocxTemplate(template_file)

img = InlineImage(doc, output_img, Cm(14))
context = { "date": "2022/07/30",
            "authors": ["陳會安", "江小魚"],
            "sales_data": sales,
            "sales_chart": img }

doc.render(context)
doc.save(output_file)
