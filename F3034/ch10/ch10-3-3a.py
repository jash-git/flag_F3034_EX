import openpyxl

# 讀取Excel檔案
wb = openpyxl.load_workbook('customer_service.xlsx')
# 選擇第一個工作表
ws = wb.active

# 使用迴圈逐一讀取C2到C5的儲存格內容
for row in ws.iter_rows(min_row=2, max_row=5, min_col=3, max_col=3):
    for cell in row:
        # 取得前10個字
        customer_issue = cell.value[:10]
        # 將前10個字填入D欄對應的儲存格
        ws.cell(row=cell.row, column=4, value=customer_issue)

# 儲存修改後的Excel檔案
wb.save('customer_service_updated.xlsx')
