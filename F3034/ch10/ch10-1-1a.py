import openpyxl

def cell2index(cell):
    """
    將儲存格標籤轉換為列和欄的索引

    Args:
        cell (str): 儲存格標籤，例如'B2'

    Returns:
        tuple: 儲存格索引，例如(2, 2)
    """
    column = ord(cell[0]) - 64
    row = int(cell[1:])
    return row, column

# 讀取Excel檔案
workbook = openpyxl.load_workbook('excel/example.xlsx')

# 選擇工作表
worksheet = workbook.active

# 取得儲存格的值
cell = 'B2'
row, column = cell2index(cell)
value = worksheet.cell(row=row, column=column).value

# 顯示儲存格的值
print(f'{cell}的值為：{value}')
# 取得儲存格的值
cell = 'A4'
row, column = cell2index(cell)
value = worksheet.cell(row=row, column=column).value

# 顯示儲存格的值
print(f'{cell}的值為：{value}')
