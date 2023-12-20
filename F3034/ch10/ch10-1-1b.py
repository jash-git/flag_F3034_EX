import openpyxl

def read_cells_range(cells_range, file_path):
    """
    讀取Excel檔案中指定儲存格範圍的值

    Args:
        cells_range (str): 儲存格範圍，例如'A1'、'B1:B3'、'A1:B3'
        file_path (str): Excel檔案路徑

    Returns:
        取決於傳入的儲存格範圍：
        - 如果傳入單個儲存格，回傳該儲存格的值。
        - 如果傳入多個儲存格，回傳包含這些儲存格值的元組。
        - 如果傳入的是矩陣形式的儲存格範圍，回傳包含巢狀元組的元組。
    """
    # 讀取Excel檔案
    workbook = openpyxl.load_workbook(file_path)

    # 選擇工作表
    worksheet = workbook.active

    # 解析儲存格範圍
    if ':' in cells_range:
        # 傳入的是儲存格範圍
        start, end = cells_range.split(':')
        start_row, start_col = cell2index(start)
        end_row, end_col = cell2index(end)
        values = []
        for row in range(start_row, end_row+1):
            row_values = []
            for col in range(start_col, end_col+1):
                row_values.append(worksheet.cell(row=row, column=col).value)
            values.append(tuple(row_values))
        if len(values) == 1:
            return values[0]
        else:
            return tuple(values)
    else:
        # 傳入的是單個儲存格
        row, col = cell2index(cells_range)
        return worksheet.cell(row=row, column=col).value


def cell2index(cell):
    """
    將儲存格標籤轉換為列和欄的索引

    Args:
        cell (str): 儲存格標籤，例如'B2'

    Returns:
        tuple: 儲存格索引，例如(2, 2)
    """
    column = ord(cell[0]) - 64
    row = int(cell[1:])
    return row, column


file_path = 'excel/example.xlsx'
print(read_cells_range('A1', file_path))
print(read_cells_range('B1:B3', file_path))
print(read_cells_range('A1:B3', file_path))